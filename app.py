"""
Agent Comptable IA - v5.0 SECURE
Traitement automatique de tickets de frais -> ecritures comptables Sage
Multi-provider : Claude -> OpenAI -> Ollama (fallback)
Securite : Auth, CSRF, Zero Data Retention, Anti-injection, Headers
"""

import os
import io
import json
import base64
import re
import time
import email
import imaplib
import smtplib
import threading
import secrets
import hashlib
import hmac
import logging
from logging.handlers import RotatingFileHandler
from functools import wraps
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timedelta
from pathlib import Path

import requests
from flask import (
    Flask, request, jsonify, render_template, send_file,
    session, redirect, url_for, abort, make_response,
    after_this_request
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.colors import red, black
from reportlab.lib.pagesizes import A4
import fitz  # PyMuPDF

app = Flask(__name__)


# ===================================================================
# LOGGING STRUCTURE
# ===================================================================

Path('logs').mkdir(exist_ok=True)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger('enop')

handler = RotatingFileHandler('logs/enop.log', maxBytes=5*1024*1024, backupCount=3)
handler.setFormatter(logging.Formatter(
    '{"time": "%(asctime)s", "level": "%(levelname)s", "msg": "%(message)s"}'
))
logger.addHandler(handler)


# ===================================================================
# CONFIGURATION
# ===================================================================

# --- Securite ---
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('FORCE_HTTPS', 'false').lower() == 'true'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=2)

# --- Identifiants login (via env, JAMAIS en dur) ---
APP_USERNAME = os.environ.get('APP_USERNAME', 'admin')
APP_PASSWORD_HASH = os.environ.get('APP_PASSWORD_HASH', '')
APP_PASSWORD_PLAIN = os.environ.get('APP_PASSWORD', 'changeme')

# --- API Keys ---
ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')
OPENAI_API_KEY = os.environ.get('OPENAI_API_KEY', '')
OLLAMA_URL = os.environ.get('OLLAMA_URL', 'http://localhost:11434')
OLLAMA_MODEL = os.environ.get('OLLAMA_MODEL', 'qwen3-vl')

# --- Retry & Rate Limiting ---
MAX_RETRIES = 3
RETRY_BASE_DELAY = 2
RATE_LIMIT_DELAY = 1.5
RATE_LIMIT_429_WAIT = 30

# --- Brute-force protection ---
LOGIN_ATTEMPTS_FILE = Path('login_attempts.json')
MAX_LOGIN_ATTEMPTS = 5
LOCKOUT_DURATION = 300  # 5 minutes

# --- Rate limiting /api/process ---
PROCESS_RATE_LIMIT = {}

# --- Webhook ---
WEBHOOK_TOKEN = os.environ.get('WEBHOOK_TOKEN', '')

# --- Dossiers (temporaires, nettoyes apres usage) ---
OUTPUT_FOLDER = Path('outputs')
OUTPUT_FOLDER.mkdir(exist_ok=True)

# --- Auto-delete : supprimer les fichiers de plus de X minutes ---
FILE_RETENTION_MINUTES = int(os.environ.get('FILE_RETENTION_MINUTES', '10'))

# --- Email (optionnel) ---
EMAIL_ADDRESS = os.environ.get('EMAIL_ADDRESS', '')
EMAIL_PASSWORD = os.environ.get('EMAIL_PASSWORD', '')
IMAP_SERVER = 'imap.gmail.com'
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465
CHECK_INTERVAL = 30

# --- Prompt comptable (externalise) ---
SYSTEM_PROMPT = Path('prompts/comptable.md').read_text(encoding='utf-8')


# ===================================================================
# SECURITE : HELPERS
# ===================================================================

def hash_password(password):
    """Hash un mot de passe avec SHA-256 + salt"""
    salt = secrets.token_hex(16)
    h = hashlib.sha256(f"{salt}{password}".encode()).hexdigest()
    return f"{salt}:{h}"


def verify_password(password, stored_hash):
    """Verifie un mot de passe contre son hash"""
    if ':' not in stored_hash:
        return False
    salt, h = stored_hash.split(':', 1)
    return hmac.compare_digest(
        hashlib.sha256(f"{salt}{password}".encode()).hexdigest(),
        h
    )


def check_password(password):
    """Verifie le mot de passe (hash ou plain selon config)"""
    if APP_PASSWORD_HASH:
        return verify_password(password, APP_PASSWORD_HASH)
    return hmac.compare_digest(password, APP_PASSWORD_PLAIN)


def load_attempts():
    """Charge les tentatives de login depuis le fichier JSON"""
    if LOGIN_ATTEMPTS_FILE.exists():
        try:
            data = json.loads(LOGIN_ATTEMPTS_FILE.read_text(encoding='utf-8'))
            for ip, val in data.items():
                if val[1]:
                    data[ip][1] = datetime.fromisoformat(val[1])
            return data
        except Exception:
            return {}
    return {}


def save_attempts(attempts):
    """Sauvegarde les tentatives de login dans le fichier JSON"""
    data = {}
    for ip, val in attempts.items():
        data[ip] = [val[0], val[1].isoformat() if val[1] else None]
    LOGIN_ATTEMPTS_FILE.write_text(json.dumps(data, indent=2), encoding='utf-8')


def is_locked_out(ip):
    """Verifie si une IP est bloquee pour trop de tentatives"""
    attempts = load_attempts()
    if ip in attempts:
        count, lockout_time = attempts[ip]
        if lockout_time and datetime.now() < lockout_time:
            return True
        if lockout_time and datetime.now() >= lockout_time:
            del attempts[ip]
            save_attempts(attempts)
            return False
    return False


def record_failed_attempt(ip):
    """Enregistre une tentative de login echouee"""
    attempts = load_attempts()
    if ip not in attempts:
        attempts[ip] = [0, None]
    attempts[ip][0] += 1
    if attempts[ip][0] >= MAX_LOGIN_ATTEMPTS:
        attempts[ip][1] = datetime.now() + timedelta(seconds=LOCKOUT_DURATION)
    save_attempts(attempts)


def clear_attempts(ip):
    """Reset les tentatives apres un login reussi"""
    attempts = load_attempts()
    if ip in attempts:
        del attempts[ip]
        save_attempts(attempts)


def login_required(f):
    """Decorateur pour proteger les routes"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('authenticated'):
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': 'Non authentifie'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def generate_csrf_token():
    """Genere un token CSRF"""
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)
    return session['csrf_token']


def validate_csrf(token):
    """Valide le token CSRF"""
    return hmac.compare_digest(
        token or '',
        session.get('csrf_token', '')
    )


def sanitize_filename(filename):
    """Nettoie un nom de fichier contre les injections path traversal"""
    filename = os.path.basename(filename)
    filename = re.sub(r'[^\w\s\-\.]', '', filename)
    filename = filename.strip('. ')
    if not filename:
        filename = 'document.pdf'
    return filename


def cleanup_old_files():
    """Supprime les fichiers de sortie de plus de FILE_RETENTION_MINUTES"""
    try:
        cutoff = datetime.now() - timedelta(minutes=FILE_RETENTION_MINUTES)
        for f in OUTPUT_FOLDER.iterdir():
            if f.is_file():
                mtime = datetime.fromtimestamp(f.stat().st_mtime)
                if mtime < cutoff:
                    f.unlink()
                    logger.info(f"[Cleanup] Supprime {f.name}")
    except Exception as e:
        logger.error(f"[Cleanup] Erreur: {e}")


def schedule_cleanup():
    """Lance le nettoyage automatique toutes les 5 minutes"""
    while True:
        time.sleep(300)
        cleanup_old_files()


# ===================================================================
# SECURITE : MIDDLEWARE
# ===================================================================

@app.before_request
def security_checks():
    """Verifications de securite avant chaque requete"""
    # CSRF sur les POST (sauf login et webhook)
    if request.method == 'POST' and request.path not in ('/login', '/api/webhook'):
        if session.get('authenticated'):
            token = (request.form.get('csrf_token') or
                     request.headers.get('X-CSRF-Token') or
                     '')
            if not validate_csrf(token):
                abort(403)


@app.after_request
def security_headers(response):
    """Headers de securite sur toutes les reponses"""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data:; "
        "connect-src 'self'"
    )
    # Pas de cache sur les fichiers sensibles
    if request.path.startswith('/api/download'):
        response.headers['Cache-Control'] = 'no-store'
    return response


# ===================================================================
# UTILITAIRES PDF
# ===================================================================

def extract_text_from_pdf(pdf_bytes):
    """Extrait le texte d'un PDF avec PyMuPDF"""
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        text = ""
        for page in doc:
            text += page.get_text()
        doc.close()
        return text.strip()
    except Exception:
        return ""


def split_pdf_pages(pdf_bytes, filename):
    """Decoupe un PDF en pages individuelles"""
    reader = PdfReader(io.BytesIO(pdf_bytes))
    pages = []
    for i, page in enumerate(reader.pages):
        writer = PdfWriter()
        writer.add_page(page)
        output = io.BytesIO()
        writer.write(output)
        output.seek(0)
        page_name = f"{Path(filename).stem}_page{i+1}.pdf"
        pages.append({
            'filename': page_name,
            'bytes': output.read(),
            'original_filename': filename
        })
    return pages


def stamp_pdf_with_s(pdf_bytes):
    """Ajoute un S rouge sur le PDF"""
    reader = PdfReader(io.BytesIO(pdf_bytes))
    writer = PdfWriter()
    for page in reader.pages:
        packet = io.BytesIO()
        w = float(page.mediabox.width)
        h = float(page.mediabox.height)
        c = canvas.Canvas(packet, pagesize=(w, h))
        c.setFont("Helvetica-Bold", 60)
        c.setFillColor(red)
        c.setFillAlpha(0.7)
        c.drawString(w - 70, h - 70, "S")
        c.save()
        packet.seek(0)
        overlay = PdfReader(packet)
        page.merge_page(overlay.pages[0])
        writer.add_page(page)
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output.read()


def merge_pdfs(pdf_list):
    """Fusionne plusieurs PDFs en un seul"""
    writer = PdfWriter()
    for pdf_bytes in pdf_list:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        for page in reader.pages:
            writer.add_page(page)
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output.read()


# ===================================================================
# PROVIDERS IA
# ===================================================================

def call_anthropic(user_content):
    """Appel Claude API"""
    if not ANTHROPIC_API_KEY:
        raise Exception("Anthropic: cle API non configuree")

    response = requests.post(
        'https://api.anthropic.com/v1/messages',
        headers={
            'Content-Type': 'application/json',
            'x-api-key': ANTHROPIC_API_KEY,
            'anthropic-version': '2023-06-01'
        },
        json={
            'model': 'claude-sonnet-4-20250514',
            'max_tokens': 4000,
            'system': SYSTEM_PROMPT,
            'messages': [{'role': 'user', 'content': user_content}]
        },
        timeout=120
    )

    if response.status_code == 200:
        return response.json()['content'][0]['text']

    error_msg = f"Anthropic HTTP {response.status_code}"
    try:
        error_detail = response.json().get('error', {}).get('message', '')
        if error_detail:
            error_msg += f" - {error_detail}"
    except Exception:
        pass
    raise Exception(error_msg)


def call_openai(user_content):
    """Appel OpenAI GPT-4o (fallback 1)"""
    if not OPENAI_API_KEY:
        raise Exception("OpenAI: cle API non configuree")

    if isinstance(user_content, list):
        messages_content = []
        for item in user_content:
            if item.get('type') == 'text':
                messages_content.append({"type": "text", "text": item['text']})
            elif item.get('type') == 'document':
                messages_content.append({
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:{item['source']['media_type']};base64,{item['source']['data']}"
                    }
                })
    else:
        messages_content = user_content

    response = requests.post(
        'https://api.openai.com/v1/chat/completions',
        headers={
            'Content-Type': 'application/json',
            'Authorization': f'Bearer {OPENAI_API_KEY}'
        },
        json={
            'model': 'gpt-4o',
            'max_tokens': 4000,
            'messages': [
                {'role': 'system', 'content': SYSTEM_PROMPT},
                {'role': 'user', 'content': messages_content}
            ]
        },
        timeout=120
    )

    if response.status_code == 200:
        return response.json()['choices'][0]['message']['content']
    raise Exception(f"OpenAI HTTP {response.status_code}")


def call_ollama(text_content):
    """Appel Ollama local (fallback 2 - texte uniquement)"""
    if not text_content or not isinstance(text_content, str):
        raise Exception("Ollama: pas de texte disponible")

    prompt = f"{SYSTEM_PROMPT}\n\nAnalyse ce ticket de frais :\n\n{text_content}"

    try:
        response = requests.post(
            f'{OLLAMA_URL}/api/generate',
            json={
                'model': OLLAMA_MODEL,
                'prompt': prompt,
                'stream': False,
                'options': {'temperature': 0.1, 'num_predict': 4000}
            },
            timeout=180
        )
    except requests.exceptions.ConnectionError:
        raise Exception("Ollama: serveur non accessible")

    if response.status_code == 200:
        return response.json().get('response', '')
    raise Exception(f"Ollama HTTP {response.status_code}")


# ===================================================================
# MOTEUR D'ANALYSE AVEC RETRY + FALLBACK
# ===================================================================

def clean_json_response(text):
    """Nettoie et parse la reponse JSON"""
    text = re.sub(r'```json\s*', '', text)
    text = re.sub(r'```\s*', '', text).strip()
    json_match = re.search(r'\{.*\}', text, re.DOTALL)
    if json_match:
        text = json_match.group()
    return json.loads(text)

def validate_and_fix_ecritures(ecritures):
    """Post-traitement Python : verifie et corrige les maths"""
    alerts = []

    for e in ecritures:
        e['debit'] = round(float(e.get('debit', 0) or 0), 2)
        e['credit'] = round(float(e.get('credit', 0) or 0), 2)

        # Comptes sur 8 caracteres
        compte = str(e.get('compte', ''))
        if len(compte) < 8:
            e['compte'] = compte.ljust(8, '0')

        # Journal toujours FCB
        e['journal'] = 'FCB'

        # Pas de montants negatifs
        if e['debit'] < 0:
            e['debit'] = abs(e['debit'])
        if e['credit'] < 0:
            e['credit'] = abs(e['credit'])

    # Verif : une seule ligne credit (banque 51200000)
    lignes_debit = [e for e in ecritures if e['debit'] > 0]
    lignes_credit = [e for e in ecritures if e['credit'] > 0]

    if lignes_debit and lignes_credit:
        total_debit = round(sum(e['debit'] for e in lignes_debit), 2)
        total_credit = round(sum(e['credit'] for e in lignes_credit), 2)

        # Forcer equilibre : credit banque = somme debits
        if abs(total_debit - total_credit) > 0.01:
            ligne_banque = next((e for e in ecritures if e['compte'] == '51200000'), None)
            if ligne_banque:
                ligne_banque['credit'] = total_debit
                alerts.append(f"Credit banque ajuste: {total_credit} -> {total_debit}")

    # Verif : ligne TVA coherente
    ligne_tva = next((e for e in ecritures if e['compte'] == '44566000'), None)
    ligne_charge = next((e for e in ecritures if e['debit'] > 0 and e['compte'] != '44566000'), None)
    ligne_banque = next((e for e in ecritures if e['compte'] == '51200000'), None)

    if ligne_tva and ligne_charge and ligne_banque:
        tva = ligne_tva['debit']
        charge = ligne_charge['debit']
        banque = ligne_banque['credit']

        # HT + TVA doit = TTC (banque)
        if abs((charge + tva) - banque) > 0.02:
            # Recalcul : charge = banque - tva
            ligne_charge['debit'] = round(banque - tva, 2)
            alerts.append(f"Charge recalculee: {charge} -> {ligne_charge['debit']}")

    # Verif finale
    total_d = round(sum(e['debit'] for e in ecritures), 2)
    total_c = round(sum(e['credit'] for e in ecritures), 2)
    if abs(total_d - total_c) > 0.01:
        ligne_banque = next((e for e in ecritures if e['compte'] == '51200000'), None)
        if ligne_banque:
            ligne_banque['credit'] = total_d
            alerts.append("Equilibre force en dernier recours")

    return ecritures, alerts

def analyze_ticket_with_retry(pdf_bytes, filename="ticket.pdf"):
    """Analyse avec fallback : Claude -> OpenAI -> Ollama"""
    text = extract_text_from_pdf(pdf_bytes)
    has_text = len(text.strip()) > 50

    if has_text:
        cloud_content = f"Analyse ce ticket de frais et produis les ecritures comptables :\n\n{text}"
    else:
        pdf_b64 = base64.b64encode(pdf_bytes).decode('utf-8')
        cloud_content = [
            {
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": pdf_b64
                }
            },
            {
                "type": "text",
                "text": "Analyse ce ticket de frais et produis les ecritures comptables. "
                        "Une page peut contenir plusieurs tickets, traite-les tous separement."
            }
        ]

    providers = []
    if ANTHROPIC_API_KEY:
        providers.append(("Claude", lambda c=cloud_content: call_anthropic(c)))
    if OPENAI_API_KEY:
        providers.append(("OpenAI", lambda c=cloud_content: call_openai(c)))
    if has_text:
        providers.append(("Ollama", lambda t=text: call_ollama(t)))

    if not providers:
        return {
            "exploitable": False,
            "raison_non_exploitable": "Aucun provider IA configure",
            "ecritures": []
        }

    last_error = ""
    for provider_name, provider_fn in providers:
        for attempt in range(MAX_RETRIES):
            try:
                logger.info(f"[{provider_name}] {filename} - tentative {attempt+1}/{MAX_RETRIES}")
                raw_response = provider_fn()
                result = clean_json_response(raw_response)
                if 'exploitable' not in result:
                    raise ValueError("JSON sans champ 'exploitable'")
                logger.info(f"[{provider_name}] {filename} - OK")
                return result

            except json.JSONDecodeError as e:
                last_error = f"{provider_name}: JSON invalide ({e})"
                logger.info(f"[{provider_name}] JSON invalide, retry...")
                time.sleep(RETRY_BASE_DELAY)

            except ValueError as e:
                last_error = f"{provider_name}: {e}"
                logger.info(f"[{provider_name}] {e}, retry...")
                time.sleep(RETRY_BASE_DELAY)

            except Exception as e:
                error_str = str(e)
                last_error = f"{provider_name}: {error_str}"
                logger.error(f"[{provider_name}] Erreur: {error_str}")

                if '429' in error_str:
                    wait = RATE_LIMIT_429_WAIT * (attempt + 1)
                    logger.info(f"[{provider_name}] Rate limit 429, attente {wait}s...")
                    time.sleep(wait)
                    continue
                if '529' in error_str:
                    wait = RETRY_BASE_DELAY * (attempt + 1) * 2
                    logger.info(f"[{provider_name}] Surcharge 529, attente {wait}s...")
                    time.sleep(wait)
                    continue
                if '400' in error_str:
                    logger.info(f"[{provider_name}] Erreur 400, provider suivant")
                    break
                time.sleep(RETRY_BASE_DELAY * (attempt + 1))

        logger.info(f"[{provider_name}] Echec apres {MAX_RETRIES} tentatives")

    return {
        "exploitable": False,
        "raison_non_exploitable": f"Analyse impossible: {last_error}",
        "ecritures": []
    }


# ===================================================================
# GENERATION EXCEL SAGE
# ===================================================================

def create_excel(all_ecritures, alerts=None, low_confidence_refs=None):
    """Cree le fichier Excel format Sage"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ecritures comptables"

    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    orange_fill = PatternFill(start_color='FFB347', end_color='FFB347', fill_type='solid')

    headers = ['Date', 'Reference', 'Journal', 'Compte', 'Libelle', 'Debit', 'Credit']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    row = 2
    total_debit = 0
    total_credit = 0

    for e in all_ecritures:
        debit = round(float(e.get('debit', 0) or 0), 2)
        credit = round(float(e.get('credit', 0) or 0), 2)
        total_debit += debit
        total_credit += credit

        is_low_confidence = (
            low_confidence_refs and e.get('reference', '') in low_confidence_refs
        )

        values = [
            e.get('date', ''),
            e.get('reference', ''),
            e.get('journal', 'FCB'),
            e.get('compte', ''),
            e.get('libelle', ''),
            debit,
            credit
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            if is_low_confidence:
                cell.fill = orange_fill
            if col in (6, 7):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
        row += 1

    row += 1
    equilibre = abs(total_debit - total_credit) < 0.01
    ctrl_fill = PatternFill(
        start_color='27AE60' if equilibre else 'E74C3C',
        end_color='27AE60' if equilibre else 'E74C3C',
        fill_type='solid'
    )
    ctrl_font = Font(name='Calibri', bold=True, color='FFFFFF')

    ws.cell(row=row, column=4, value='CONTROLE').font = ctrl_font
    ws.cell(row=row, column=4).fill = ctrl_fill
    status = 'OK - Equilibre' if equilibre else 'ERREUR - Desequilibre'
    ws.cell(row=row, column=5, value=status).font = ctrl_font
    ws.cell(row=row, column=5).fill = ctrl_fill
    ws.cell(row=row, column=6, value=round(total_debit, 2)).font = ctrl_font
    ws.cell(row=row, column=6).fill = ctrl_fill
    ws.cell(row=row, column=6).number_format = '#,##0.00'
    ws.cell(row=row, column=7, value=round(total_credit, 2)).font = ctrl_font
    ws.cell(row=row, column=7).fill = ctrl_fill
    ws.cell(row=row, column=7).number_format = '#,##0.00'

    if alerts:
        row += 2
        alert_font = Font(name='Calibri', bold=True, color='E74C3C')
        ws.cell(row=row, column=1, value='ALERTES').font = alert_font
        for alert in alerts:
            row += 1
            ws.cell(row=row, column=1, value=alert)

    for col_letter, width in [('A', 14), ('B', 12), ('C', 10), ('D', 12), ('E', 45), ('F', 14), ('G', 14)]:
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ===================================================================
# RAPPORT INEXPLOITABLES
# ===================================================================

def create_inexploitable_report(inexploitable_tickets):
    """Cree un PDF listant les tickets inexploitables"""
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    c.setFont("Helvetica-Bold", 18)
    c.setFillColor(red)
    c.drawString(50, height - 60, "Justificatifs inexploitables")
    c.setFont("Helvetica", 11)
    c.setFillColor(black)
    c.drawString(50, height - 85, f"Date de traitement : {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.setFont("Helvetica", 10)
    c.drawString(50, height - 110, "Les documents suivants n'ont pas pu etre exploites.")
    c.drawString(50, height - 125, "Merci de fournir des justificatifs conformes.")

    y = height - 165
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, y, "Fichier")
    c.drawString(300, y, "Motif")
    y -= 5
    c.line(50, y, width - 50, y)
    y -= 20

    c.setFont("Helvetica", 10)
    for ticket in inexploitable_tickets:
        if y < 80:
            c.showPage()
            y = height - 60
        c.drawString(50, y, ticket['filename'][:35])
        c.drawString(300, y, ticket['raison'][:50])
        y -= 18

    c.save()
    buffer.seek(0)
    return buffer.read()


# ===================================================================
# TRAITEMENT PRINCIPAL
# ===================================================================

def process_tickets(files_data):
    """Traite une liste de tickets"""
    all_ecritures = []
    exploited_pdfs = []
    inexploitable_tickets = []
    alerts = []
    low_confidence_refs = set()
    ticket_num = 1
    results_detail = []

    # Split multi-pages (seulement si >20 pages)
    split_files = []
    for file_info in files_data:
        try:
            reader = PdfReader(io.BytesIO(file_info['bytes']))
            if len(reader.pages) > 20:
                logger.info(f"Split {file_info['filename']} : {len(reader.pages)} pages")
                pages = split_pdf_pages(file_info['bytes'], file_info['filename'])
                split_files.extend(pages)
            else:
                split_files.append(file_info)
        except Exception as e:
            logger.error(f"Erreur split {file_info['filename']}: {e}")
            split_files.append(file_info)

    total_pages = len(split_files)
    logger.info(f"{'='*50}")
    logger.info(f"Traitement de {total_pages} page(s)")
    logger.info(f"{'='*50}")

    for idx, file_info in enumerate(split_files):
        filename = file_info['filename']
        pdf_bytes = file_info['bytes']

        logger.info(f"[{idx+1}/{total_pages}] {filename}")
        result = analyze_ticket_with_retry(pdf_bytes, filename)

        # Verification confiance
        if result.get('confidence', 1.0) < 0.7:
            alerts.append(
                f"\u26a0\ufe0f Confiance faible ({result['confidence']:.0%}) "
                f"sur {filename} \u2014 verification manuelle recommandee"
            )

        if result.get('exploitable'):
            ecritures = result.get('ecritures', [])
            for e in ecritures:
                e['reference'] = f'T{ticket_num}'

            # Tracker les references a faible confiance
            if result.get('confidence', 1.0) < 0.7:
                low_confidence_refs.add(f'T{ticket_num}')

            # Post-traitement Python
            ecritures, fix_alerts = validate_and_fix_ecritures(ecritures)
            for a in fix_alerts:
                alerts.append(f"T{ticket_num} ({filename}) : {a}")

            # Verification equilibre
            total_d = sum(e['debit'] for e in ecritures)
            total_c = sum(e['credit'] for e in ecritures)
            if abs(total_d - total_c) > 0.01:
                alerts.append(f"T{ticket_num} ({filename}) : Desequilibre ({total_d:.2f} != {total_c:.2f})")
                ligne_banque = next((e for e in ecritures if e['compte'] == '51200000'), None)
                if ligne_banque:
                    ligne_banque['credit'] = round(total_d, 2)

            all_ecritures.extend(ecritures)
            stamped = stamp_pdf_with_s(pdf_bytes)
            exploited_pdfs.append(stamped)
            results_detail.append({
                'filename': filename, 'status': 'exploitable',
                'reference': f'T{ticket_num}', 'ecritures': ecritures
            })
            ticket_num += 1
        else:
            raison = result.get('raison_non_exploitable', 'Document inexploitable')
            inexploitable_tickets.append({'filename': filename, 'raison': raison})
            alerts.append(f"!! {filename} : {raison}")
            results_detail.append({
                'filename': filename, 'status': 'inexploitable', 'raison': raison
            })

        if idx < total_pages - 1:
            time.sleep(RATE_LIMIT_DELAY)

    # Generation fichiers (supprimes automatiquement apres FILE_RETENTION_MINUTES)
    output_files = {}
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    if all_ecritures:
        excel_bytes = create_excel(
            all_ecritures,
            alerts if alerts else None,
            low_confidence_refs=low_confidence_refs
        )
        excel_name = f'Sage_import_{timestamp}.xlsx'
        (OUTPUT_FOLDER / excel_name).write_bytes(excel_bytes)
        output_files['excel'] = {'name': excel_name, 'path': str(OUTPUT_FOLDER / excel_name)}

    if exploited_pdfs:
        merged = merge_pdfs(exploited_pdfs)
        stamped_name = f'Tickets_exploites_S_{timestamp}.pdf'
        (OUTPUT_FOLDER / stamped_name).write_bytes(merged)
        output_files['stamped_pdf'] = {'name': stamped_name, 'path': str(OUTPUT_FOLDER / stamped_name)}

    if inexploitable_tickets:
        report = create_inexploitable_report(inexploitable_tickets)
        report_name = f'Justificatifs_inexploites_{timestamp}.pdf'
        (OUTPUT_FOLDER / report_name).write_bytes(report)
        output_files['inexploitable_pdf'] = {'name': report_name, 'path': str(OUTPUT_FOLDER / report_name)}

    total_d = round(sum(e['debit'] for e in all_ecritures), 2)
    total_c = round(sum(e['credit'] for e in all_ecritures), 2)
    logger.info(f"{'='*50}")
    logger.info(f"RESULTAT : {len(exploited_pdfs)} exploites / {len(inexploitable_tickets)} inexploitables")
    logger.info(f"TOTAUX   : D={total_d} | C={total_c} | {'OK' if abs(total_d - total_c) < 0.01 else 'ERREUR'}")
    logger.info(f"{'='*50}")

    return {
        'output_files': output_files,
        'results_detail': results_detail,
        'summary': {
            'total': total_pages,
            'exploites': len(exploited_pdfs),
            'inexploites': len(inexploitable_tickets),
            'total_debit': total_d,
            'total_credit': total_c,
            'equilibre': abs(total_d - total_c) < 0.01
        }
    }


# ===================================================================
# EMAIL
# ===================================================================

def send_email_with_attachments(to_email, subject, body, attachments):
    msg = MIMEMultipart()
    msg['From'] = EMAIL_ADDRESS
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))
    for att_filename, file_bytes in attachments:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(file_bytes)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{att_filename}"')
        msg.attach(part)
    with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
        server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        server.send_message(msg)


def check_emails_once():
    """Une iteration de verification des emails"""
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
    mail.select('INBOX')
    _, messages = mail.search(None, 'UNSEEN')
    for num in messages[0].split():
        if not num:
            continue
        _, msg_data = mail.fetch(num, '(RFC822)')
        msg = email.message_from_bytes(msg_data[0][1])
        sender = email.utils.parseaddr(msg['From'])[1]
        subject = msg['Subject'] or 'Sans objet'
        files_data = []
        for part in msg.walk():
            if part.get_content_type() == 'application/pdf':
                att_filename = part.get_filename() or 'document.pdf'
                pdf_bytes = part.get_payload(decode=True)
                if pdf_bytes:
                    files_data.append({'filename': att_filename, 'bytes': pdf_bytes})
        if not files_data:
            continue
        logger.info(f"[EMAIL] Mail de {sender} - {len(files_data)} PDF(s)")
        results = process_tickets(files_data)
        attachments = []
        files = results['output_files']
        for key in ['excel', 'stamped_pdf', 'inexploitable_pdf']:
            if files.get(key):
                with open(files[key]['path'], 'rb') as f:
                    attachments.append((files[key]['name'], f.read()))
        s = results['summary']
        body = f"""Bonjour,

Traitement de vos {s['total']} justificatif(s) termine.

- {s['exploites']} exploite(s)
- {s['inexploites']} inexploitable(s)
- Debit : {s['total_debit']:.2f} EUR
- Credit : {s['total_credit']:.2f} EUR
- Equilibre : {'OK' if s['equilibre'] else 'ERREUR'}

Agent Comptable IA"""
        send_email_with_attachments(sender, f"Re: {subject}", body, attachments)
        logger.info(f"[EMAIL] Reponse envoyee a {sender}")
    mail.logout()


def check_emails():
    """Boucle watchdog pour la verification des emails"""
    while True:
        try:
            check_emails_once()
        except Exception as e:
            logger.error(f"[EMAIL] Crash recupere, redemarrage dans 60s: {e}")
            time.sleep(60)
            continue
        time.sleep(CHECK_INTERVAL)


# ===================================================================
# ROUTES
# ===================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        ip = request.remote_addr

        if is_locked_out(ip):
            error = "Trop de tentatives. Reessayez dans 5 minutes."
        else:
            username = request.form.get('username', '').strip()
            password = request.form.get('password', '')

            if username == APP_USERNAME and check_password(password):
                session.permanent = True
                session['authenticated'] = True
                session['login_time'] = datetime.now().isoformat()
                session['csrf_token'] = secrets.token_hex(32)
                clear_attempts(ip)
                return redirect(url_for('index'))
            else:
                record_failed_attempt(ip)
                attempts = load_attempts()
                remaining = MAX_LOGIN_ATTEMPTS - attempts.get(ip, [0, None])[0]
                if remaining > 0:
                    error = f"Identifiants incorrects. {remaining} tentative(s) restante(s)."
                else:
                    error = "Compte bloque pour 5 minutes."

    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    return render_template('index.html', csrf_token=generate_csrf_token())


@app.route('/api/process', methods=['POST'])
@login_required
def api_process():
    # Rate limiting : max 10 appels par session par heure
    rate_key = f"{session.get('login_time', '')}_{request.remote_addr}"
    now = time.time()
    if rate_key not in PROCESS_RATE_LIMIT:
        PROCESS_RATE_LIMIT[rate_key] = []
    PROCESS_RATE_LIMIT[rate_key] = [t for t in PROCESS_RATE_LIMIT[rate_key] if now - t < 3600]
    if len(PROCESS_RATE_LIMIT[rate_key]) >= 10:
        return jsonify({'error': 'Trop de requetes, attendez avant de resoumettre'}), 429
    PROCESS_RATE_LIMIT[rate_key].append(now)

    if 'files' not in request.files:
        return jsonify({'error': 'Aucun fichier envoye'}), 400

    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': 'Aucun fichier selectionne'}), 400

    files_data = []
    for f in files:
        if f.filename and f.filename.lower().endswith('.pdf'):
            safe_name = sanitize_filename(f.filename)
            pdf_bytes = f.read()

            # Validation : verifier que c'est bien un PDF
            if not pdf_bytes[:5] == b'%PDF-':
                continue

            files_data.append({'filename': safe_name, 'bytes': pdf_bytes})

    if not files_data:
        return jsonify({'error': 'Aucun fichier PDF valide'}), 400

    try:
        results = process_tickets(files_data)

        # Nettoyage immediat des donnees en memoire
        for fd in files_data:
            fd['bytes'] = None
        files_data.clear()

        return jsonify(results)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/download/<filename>')
@login_required
def download_file(filename):
    # Anti path-traversal
    safe_name = sanitize_filename(filename)
    filepath = OUTPUT_FOLDER / safe_name

    if not filepath.exists():
        return jsonify({'error': 'Fichier non trouve'}), 404

    # Verifier que le fichier est bien dans OUTPUT_FOLDER
    try:
        filepath.resolve().relative_to(OUTPUT_FOLDER.resolve())
    except ValueError:
        abort(403)

    @after_this_request
    def remove_file(response):
        try:
            filepath.unlink()
            logger.info(f"[ZDR] Fichier supprime apres download: {safe_name}")
        except Exception:
            pass
        return response

    return send_file(filepath, as_attachment=True, download_name=safe_name)


@app.route('/api/status')
@login_required
def api_status():
    providers = {
        'anthropic': bool(ANTHROPIC_API_KEY),
        'openai': bool(OPENAI_API_KEY),
        'ollama': False
    }
    try:
        r = requests.get(f'{OLLAMA_URL}/api/tags', timeout=3)
        providers['ollama'] = r.status_code == 200
    except Exception:
        pass

    return jsonify({
        'providers': providers,
        'active_providers': sum(1 for v in providers.values() if v),
        'file_retention_minutes': FILE_RETENTION_MINUTES
    })


@app.route('/api/webhook', methods=['POST'])
def webhook():
    """Endpoint webhook pour OpenClaw"""
    # Auth par token Bearer (separe du systeme de session web)
    auth_header = request.headers.get('Authorization', '')
    webhook_token = os.environ.get('WEBHOOK_TOKEN', '')

    if not webhook_token or auth_header != f'Bearer {webhook_token}':
        return jsonify({'error': 'Non autorise'}), 401

    # Accepte JSON avec base64 des PDFs
    data = request.get_json()
    if not data or 'files' not in data:
        return jsonify({'error': 'Format invalide, attendu: {"files": [{"name": "...", "data": "base64..."}]}'}), 400

    files_data = []
    for f in data['files']:
        pdf_bytes = base64.b64decode(f['data'])
        if pdf_bytes[:5] != b'%PDF-':
            continue
        files_data.append({
            'filename': sanitize_filename(f.get('name', 'document.pdf')),
            'bytes': pdf_bytes
        })

    if not files_data:
        return jsonify({'error': 'Aucun PDF valide'}), 400

    results = process_tickets(files_data)

    # Retourne le summary + les fichiers en base64
    response_data = {'summary': results['summary'], 'files': {}}
    for key, file_info in results['output_files'].items():
        with open(file_info['path'], 'rb') as fh:
            response_data['files'][key] = {
                'name': file_info['name'],
                'data': base64.b64encode(fh.read()).decode()
            }

    return jsonify(response_data)


# ===================================================================
# DEMARRAGE
# ====================================================================

if __name__ == '__main__':
    logger.info("=" * 50)
    logger.info("  AGENT COMPTABLE IA v5.0 SECURE")
    logger.info("=" * 50)

    logger.info("Securite :")
    logger.info(f"  Login         : {APP_USERNAME} / {'hash' if APP_PASSWORD_HASH else 'plain'}")
    logger.info(f"  Session       : {app.config['PERMANENT_SESSION_LIFETIME']}")
    logger.info(f"  CSRF          : actif")
    logger.info(f"  Anti-bruteforce: {MAX_LOGIN_ATTEMPTS} tentatives, lockout {LOCKOUT_DURATION}s")
    logger.info(f"  Zero Data     : fichiers supprimes apres {FILE_RETENTION_MINUTES} min")
    logger.info(f"  Headers       : CSP, X-Frame-Options, nosniff, no-cache")

    logger.info("Providers :")
    logger.info(f"  Claude  : {'OK' if ANTHROPIC_API_KEY else 'NON'}")
    logger.info(f"  OpenAI  : {'OK' if OPENAI_API_KEY else 'NON'}")
    try:
        r = requests.get(f'{OLLAMA_URL}/api/tags', timeout=3)
        logger.info(f"  Ollama  : {'OK - ' + OLLAMA_MODEL if r.status_code == 200 else 'NON'}")
    except Exception:
        logger.info(f"  Ollama  : NON ({OLLAMA_URL})")

    logger.info(f"  Webhook : {'actif sur /api/webhook' if WEBHOOK_TOKEN else 'desactive (WEBHOOK_TOKEN non defini)'}")

    if EMAIL_ADDRESS and EMAIL_PASSWORD:
        email_thread = threading.Thread(target=check_emails, daemon=True)
        email_thread.start()
        logger.info(f"Email : {EMAIL_ADDRESS}")
    else:
        logger.info("Email : non configure")

    # Thread de nettoyage automatique
    cleanup_thread = threading.Thread(target=schedule_cleanup, daemon=True)
    cleanup_thread.start()

    logger.info(f"Interface : http://localhost:5000")
    logger.info("=" * 50)

    app.run(host='0.0.0.0', port=5000, debug=False)